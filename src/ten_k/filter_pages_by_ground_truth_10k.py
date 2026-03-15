"""Filter PDF pages that contain financial statements matching the ground truth data.

Scans each 10-K PDF for pages containing the three core financial statements
(Income Statement, Balance Sheet, Cash Flow Statement) by combining:
  1. Header detection — looks for statement titles near the top of the page.
  2. Value matching — checks whether distinctive ground-truth numbers appear on the page,
     accounting for unit differences (thousands vs millions) and formatting (commas, parens).

Outputs a JSON mapping of each PDF to its matched pages, and optionally extracts
those pages into separate PDFs.
"""

import argparse
import json
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pymupdf

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SHEET_TO_PDF: dict[str, str] = {
    "20241231_AMER_PUB_NEWELL": "AMER_PUB_NEWELL_Q4_Annual_FS.pdf",
    "20243112_AMER_PUB_MARAVAI": "AMER_PUB_MARAVAI_Q4_Annual_FS.pdf",
    "20243112_AMER_PUB_BELLRING": "AMER_PUB_BELLRING_Q4_Annual_FS.pdf",
    "20241231_AMER_PUB_DOORDASH": "AMER_PUB_DOORDASH_Q4_Annual_FS.pdf",
    "20241231_AMER_PUB_FloorandDecor": "AMER_PUB_FloorandDecor_Q4_Annual_FS.pdf",
}

# Header patterns per statement type (checked case-insensitively)
HEADER_PATTERNS: dict[str, list[str]] = {
    "income_statement": [
        "consolidated statements of operations",
        "consolidated statements of income",
        "consolidated statements of earnings",
        "consolidated statement of operations",
        "consolidated statement of income",
    ],
    "balance_sheet": [
        "consolidated balance sheet",
    ],
    "cash_flow": [
        "consolidated statements of cash flow",
        "consolidated statement of cash flow",
    ],
}

# Ground-truth line items mapped to each statement type
STATEMENT_ITEMS: dict[str, list[str]] = {
    "income_statement": [
        "Revenue",
        "COGS",
        "Gross Profit",
        "SG&A",
        "Total Operating Expenses",
        "Net Income",
        "Plus: Taxes",
        "Plus: Interest Expense",
        "Plus: D&A",
    ],
    "balance_sheet": [
        "Cash",
        "Accounts Receivable",
        "Inventory",
        "Current Assets",
        "Goodwill",
        "Other Intangibles",
        "Total Assets",
        "Short Term Debt",
        "Accounts Payable",
        "Accrued Expenses",
        "Deferred Revenue",
        "Current Liabilities",
        "Total Liabilities",
        "Shareholders Equity",
        "Operating Lease Obligations",
    ],
    "cash_flow": [
        "Cash from Operations",
        "Cash from Investing",
        "Cash from Financing",
        "Effect of Exchange Rates / Other",
        "Change in Cash",
        "Less: Cash interest (net)",
        "Less: Cash taxes",
        "Less: Capex",
        "Less: Changes in NWC",
        "Acquisitions",
        "Divestitures",
        "Dividends",
        "Net Share Issuance (Repurchase)",
        "Net Debt Issuance (Repayment)",
    ],
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def number_representations(val: float | int) -> list[str]:
    """Return plausible string representations of a numeric value as it may appear in a PDF.

    Skips zero values (they match too broadly).  Generates representations
    for the value as-is and also scaled by ×1000 (to handle ground-truth in
    millions vs PDF reporting in thousands).
    """
    if val is None or val == 0:
        return []

    reps: set[str] = set()
    for multiplier in (1, 1000):
        raw = abs(val) * multiplier
        # Skip very small values after multiplier — they are not distinctive
        if raw < 1:
            continue
        int_val = int(round(raw))
        # Integer forms
        if abs(raw - int_val) < 0.5:
            reps.add(str(int_val))
            reps.add(f"{int_val:,}")
        else:
            # Decimal forms
            for fmt in ("{:.1f}", "{:.3f}"):
                plain = fmt.format(raw)
                reps.add(plain)
            if raw >= 1_000:
                for fmt in ("{:,.1f}", "{:,.3f}"):
                    reps.add(fmt.format(raw))
    return list(reps)


@dataclass
class PageMatch:
    page_num: int  # 1-indexed
    statement_type: str
    score: float
    matched_items: list[str] = field(default_factory=list)


def score_page(
    text: str,
    statement_type: str,
    gt_values: dict[str, float | int],
) -> PageMatch | None:
    """Score a page for how well it matches a given statement type."""
    text_lower = text.lower()
    headers = HEADER_PATTERNS[statement_type]
    items = STATEMENT_ITEMS[statement_type]

    # Header scoring
    header_near_top = any(h in text_lower[:400] for h in headers)
    header_anywhere = any(h in text_lower for h in headers)

    # Value matching — only non-zero values for the statement type
    matched: list[str] = []
    for item_name in items:
        if item_name not in gt_values:
            continue
        val = gt_values[item_name]
        for rep in number_representations(val):
            if rep in text:
                matched.append(item_name)
                break

    header_score = 10 if header_near_top else (3 if header_anywhere else 0)
    score = header_score + len(matched)

    # Require a meaningful match: at least a header OR strong value evidence
    if header_near_top and len(matched) >= 2:
        return PageMatch(page_num=0, statement_type=statement_type, score=score, matched_items=matched)
    if len(matched) >= 5 and header_anywhere:
        return PageMatch(page_num=0, statement_type=statement_type, score=score, matched_items=matched)
    return None


def find_financial_pages(
    pdf_path: str | Path,
    gt_values: dict[str, float | int],
) -> dict[str, PageMatch]:
    """Return the best-matching page for each statement type in a PDF."""
    doc = pymupdf.open(str(pdf_path))
    best: dict[str, PageMatch] = {}

    for i in range(len(doc)):
        text = doc[i].get_text()
        for stype in STATEMENT_ITEMS:
            match = score_page(text, stype, gt_values)
            if match is None:
                continue
            match.page_num = i + 1
            if stype not in best or match.score > best[stype].score:
                best[stype] = match

    doc.close()
    return best


def load_ground_truth(xlsx_path: str | Path) -> dict[str, dict[str, float | int]]:
    """Load ground truth values from each sheet. Returns {sheet_name: {item: value}}."""
    wb = openpyxl.load_workbook(str(xlsx_path))
    result: dict[str, dict[str, float | int]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        gt: dict[str, float | int] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            item_name = row[0]
            val = next((v for v in row[1:] if v is not None), None)
            if val is not None:
                gt[item_name] = val
        result[sheet_name] = gt
    return result


def extract_pages(pdf_path: str | Path, pages: list[int], output_path: str | Path) -> None:
    """Extract specific 1-indexed pages from a PDF into a new PDF."""
    doc = pymupdf.open(str(pdf_path))
    new_doc = pymupdf.open()
    for p in sorted(pages):
        new_doc.insert_pdf(doc, from_page=p - 1, to_page=p - 1)
    new_doc.save(str(output_path))
    new_doc.close()
    doc.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--data-dir",
        default="10K-10Q-files/Data",
        help="Directory containing the 10-K PDF files",
    )
    parser.add_argument(
        "--ground-truth",
        default="10K-10Q-files/Ground_truth_or_Source_of_Truth/10K_source_of_truth_MS_Hackathon.xlsx",
        help="Path to the ground-truth Excel file",
    )
    parser.add_argument(
        "--output-dir",
        default="data/k10/raw_documents",
        help="Directory to write extracted filtered PDFs",
    )
    parser.add_argument(
        "--eval-dataset",
        default="data/k10/evaluation_dataset.json",
        help="Path to the evaluation dataset JSON to update with document links",
    )
    parser.add_argument(
        "--json-output",
        default=None,
        help="Path to write the JSON results mapping",
    )
    args = parser.parse_args()

    ground_truth = load_ground_truth(args.ground_truth)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    results: dict[str, dict] = {}

    for sheet_name, gt_values in ground_truth.items():
        pdf_file = SHEET_TO_PDF.get(sheet_name)
        if not pdf_file:
            print(f"WARNING: No PDF mapping for sheet '{sheet_name}'")
            continue

        pdf_path = Path(args.data_dir) / pdf_file
        if not pdf_path.exists():
            print(f"WARNING: PDF not found: {pdf_path}")
            continue

        best = find_financial_pages(pdf_path, gt_values)

        print(f"\n{'=' * 70}")
        print(f"  {sheet_name} -> {pdf_file}")
        print(f"{'=' * 70}")

        pdf_result: dict[str, dict] = {"pdf_file": pdf_file, "statements": {}}
        all_pages: list[int] = []

        for stype in ("income_statement", "balance_sheet", "cash_flow"):
            if stype in best:
                m = best[stype]
                print(f"  {stype:20s}: Page {m.page_num:3d}  "
                      f"(score={m.score:.0f}, {len(m.matched_items)} values matched)")
                print(f"    Matched items: {m.matched_items}")
                pdf_result["statements"][stype] = {
                    "page": m.page_num,
                    "score": m.score,
                    "matched_items": m.matched_items,
                }
                all_pages.append(m.page_num)
            else:
                print(f"  {stype:20s}: NOT FOUND")

        pdf_result["filtered_pages"] = sorted(set(all_pages))

        # Extract matched pages into a filtered PDF
        if all_pages:
            out_filename = f"filtered_{pdf_file}"
            out_path = out_dir / out_filename
            extract_pages(pdf_path, sorted(set(all_pages)), out_path)
            pdf_result["document"] = str(out_dir / out_filename)
            print(f"  -> Extracted {len(set(all_pages))} pages to {out_path}")

        results[sheet_name] = pdf_result

    # Summary
    print(f"\n{'=' * 70}")
    print("  SUMMARY")
    print(f"{'=' * 70}")
    for sheet_name, res in results.items():
        pages = res["filtered_pages"]
        print(f"  {res['pdf_file']:50s} -> pages {pages}")

    # Update evaluation_dataset.json with document links
    eval_path = Path(args.eval_dataset)
    if eval_path.exists():
        with open(eval_path) as f:
            eval_data = json.load(f)

        for entry in eval_data:
            sheet_name = entry.get("sheet_name")
            if sheet_name in results and "document" in results[sheet_name]:
                entry["document"] = results[sheet_name]["document"]
                entry["filtered_pages"] = results[sheet_name]["filtered_pages"]
                entry["source_pdf"] = results[sheet_name]["pdf_file"]

        with open(eval_path, "w") as f:
            json.dump(eval_data, f, indent=2)
            f.write("\n")
        print(f"\nUpdated {eval_path} with document links")
    else:
        print(f"\nWARNING: Eval dataset not found at {eval_path}, skipping update")

    # Write JSON output
    if args.json_output:
        with open(args.json_output, "w") as f:
            json.dump(results, f, indent=2)
        print(f"JSON results written to {args.json_output}")


if __name__ == "__main__":
    main()
