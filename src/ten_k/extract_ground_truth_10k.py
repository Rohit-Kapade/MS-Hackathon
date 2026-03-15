"""Extract ground truth data from Excel sheets into evaluation_dataset.json."""

import json
import re
from pathlib import Path

import openpyxl

EXCEL_PATH = Path("10K-10Q-files/Ground_truth_or_Source_of_Truth/10K_source_of_truth_MS_Hackathon.xlsx")
OUTPUT_PATH = Path("data/k10/evaluation_dataset.json")

# Mapping from ground-truth field names → (section, group)
# Based on the extraction requirement tables.
FIELD_GROUPS: dict[str, tuple[str, str]] = {
    # Income & Operation → Statement
    "Revenue":                          ("income_statement", "statement"),
    "COGS":                             ("income_statement", "statement"),
    "Gross Profit":                     ("income_statement", "statement"),
    "SG&A":                             ("income_statement", "statement"),
    "Plus: Taxes":                      ("income_statement", "statement"),
    "Plus: Interest Expense":           ("income_statement", "statement"),
    "Less: Interest Income (if available)": ("income_statement", "statement"),
    "Plus: D&A":                        ("income_statement", "statement"),
    "Total Operating Expenses":         ("income_statement", "statement"),
    # Cash Flow → Operating Activities
    "Net Income":                       ("cash_flow", "operating_activities"),
    "Cash from Operations":             ("cash_flow", "operating_activities"),
    "Less: Changes in NWC":             ("cash_flow", "operating_activities"),
    "Change in Cash":                   ("cash_flow", "operating_activities"),
    # Cash Flow → Investing Activities
    "Acquisitions":                     ("cash_flow", "investing_activities"),
    "Less: Capex":                      ("cash_flow", "investing_activities"),
    "Cash from Investing":              ("cash_flow", "investing_activities"),
    "Divestitures":                     ("cash_flow", "investing_activities"),
    # Cash Flow → Financing Activities
    "Less: Cash interest (net)":        ("cash_flow", "financing_activities"),
    "Less: Cash taxes":                 ("cash_flow", "financing_activities"),
    "Dividends":                        ("cash_flow", "financing_activities"),
    "Cash from Financing":              ("cash_flow", "financing_activities"),
    "Effect of Exchange Rates / Other": ("cash_flow", "financing_activities"),
    "Net Debt Issuance (Repayment)":    ("cash_flow", "financing_activities"),
    "Net Share Issuance (Repurchase)":  ("cash_flow", "financing_activities"),
    # Balance Sheet → Asset
    "Cash":                             ("balance_sheet", "asset"),
    "Accounts Receivable":              ("balance_sheet", "asset"),
    "Inventory":                        ("balance_sheet", "asset"),
    "Current Assets":                   ("balance_sheet", "asset"),
    "Goodwill":                         ("balance_sheet", "asset"),
    "Other Intangibles":                ("balance_sheet", "asset"),
    "Total Assets":                     ("balance_sheet", "asset"),
    # Balance Sheet → Liability
    "Short Term Debt":                  ("balance_sheet", "liability"),
    "Accounts Payable":                 ("balance_sheet", "liability"),
    "Accrued Expenses":                 ("balance_sheet", "liability"),
    "Deferred Revenue":                 ("balance_sheet", "liability"),
    "Current Liabilities":              ("balance_sheet", "liability"),
    "Total Liabilities":                ("balance_sheet", "liability"),
    "Shareholders Equity":              ("balance_sheet", "liability"),
    "Operating Lease Obligations":      ("balance_sheet", "liability"),
}

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

dataset = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

    # Extract year from the first value header (e.g. "Year 2024" -> 2024)
    year_match = re.search(r"\d{4}", headers[1])
    year = int(year_match.group()) if year_match else None

    # Build grouped ground truth: section → group → {field: value}
    grouped: dict[str, dict[str, dict[str, float | None]]] = {}
    for row in rows[1:]:
        item_name = row[0]
        if item_name is None:
            continue
        name = str(item_name)
        # Merge all value columns: take the last non-null value
        merged_val = None
        for col_idx in range(1, len(row)):
            val = row[col_idx]
            if val is not None:
                merged_val = float(val)

        if name not in FIELD_GROUPS:
            print(f"  WARNING: unmapped field '{name}' in sheet '{sheet_name}'")
            continue

        section, group = FIELD_GROUPS[name]
        grouped.setdefault(section, {}).setdefault(group, {})[name] = merged_val

    dataset.append({
        "sheet_name": sheet_name,
        "year": year,
        "ground_truth": grouped,
    })

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH.write_text(json.dumps(dataset, indent=2))

# --- Sanity checks ---
print(f"Wrote {len(dataset)} sheets to {OUTPUT_PATH}")

# Collect union of all fields seen across all sheets
all_seen_fields: set[str] = set()
for entry in dataset:
    for groups in entry["ground_truth"].values():
        for fields in groups.values():
            all_seen_fields.update(fields.keys())

# Normalize: ensure every sheet has all 39 fields (fill missing with None)
for entry in dataset:
    for section, group_map in FIELD_GROUPS.items():
        pass  # just iterating for the next loop
for entry in dataset:
    for field_name, (section, group) in FIELD_GROUPS.items():
        gt = entry["ground_truth"]
        gt.setdefault(section, {}).setdefault(group, {}).setdefault(field_name, None)

# Re-write after normalization
OUTPUT_PATH.write_text(json.dumps(dataset, indent=2))

for entry in dataset:
    total = sum(len(fields) for groups in entry["ground_truth"].values() for fields in groups.values())
    print(f"  {entry['sheet_name']}: {total} line items across {list(entry['ground_truth'].keys())}")

# Check all Excel fields are mapped
all_excel_fields: set[str] = set()
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            all_excel_fields.add(str(row[0]))

unmapped = all_excel_fields - set(FIELD_GROUPS.keys())
unused = set(FIELD_GROUPS.keys()) - all_excel_fields
if unmapped:
    print(f"\n  UNMAPPED fields (in Excel but not in FIELD_GROUPS): {unmapped}")
if unused:
    print(f"\n  UNUSED mappings (in FIELD_GROUPS but not in Excel): {unused}")
if not unmapped and not unused:
    print("\n  All field names match perfectly.")

# Check union is exactly 39 fields
assert len(all_seen_fields) == 39, (
    f"Expected 39 unique fields in union, got {len(all_seen_fields)}. "
    f"Missing: {set(FIELD_GROUPS.keys()) - all_seen_fields}, "
    f"Extra: {all_seen_fields - set(FIELD_GROUPS.keys())}"
)
print(f"  Union of all fields: {len(all_seen_fields)} (expected 39) ✓")

# Check every sheet now has all 39 fields after normalization
for entry in dataset:
    sheet_fields = set()
    for groups in entry["ground_truth"].values():
        for fields in groups.values():
            sheet_fields.update(fields.keys())
    assert len(sheet_fields) == 39, (
        f"Sheet {entry['sheet_name']} has {len(sheet_fields)} fields after normalization, expected 39"
    )
print(f"  All sheets normalized to 39 fields ✓")
